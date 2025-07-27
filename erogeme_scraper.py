# erogeme_scraper.py
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook

# 設定
url_file = "urls.txt"
output_file = "erogame_stats.xlsx"

# URL一覧読み込み
with open(url_file, "r", encoding="utf-8") as f:
    urls = [line.strip() for line in f if line.strip()]

# 結果リスト
results = []

# 各URLのデータを収集
for url in urls:
    try:
        res = requests.get(url, timeout=10)
        res.encoding = res.apparent_encoding
        soup = BeautifulSoup(res.text, "html.parser")

        title_tag = soup.select_one("div#soft-title span.bold")
        title = title_tag.text.strip() if title_tag else "タイトル不明"

        brand_tag = soup.select_one("tr#brand td a")
        brand = brand_tag.text.strip() if brand_tag else "ブランド不明"

        median = average = max_score = min_score = None

        for row in soup.select("tr"):
            th = row.find("th")
            td = row.find("td")
            if not th or not td:
                continue
            key = th.text.strip()
            val = td.text.strip()
            if key == "中央値":
                median = int(val)
            elif key == "平均値":
                average = int(val)
            elif key == "最高点":
                max_score = int(val)
            elif key == "最低点":
                min_score = int(val)

        results.append({
            "取得日時": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "タイトル": title,
            "ブランド": brand,
            "中央値": median,
            "平均値": average,
            "最高点": max_score,
            "最低点": min_score,
            "URL": url
        })

        print(f"✅ 取得成功：{title}")

    except Exception as e:
        print(f"❌ エラー：{url}")
        print(e)

# 統計処理
medians = [r["中央値"] for r in results if r["中央値"] is not None]
mean_median = sum(medians) / len(medians)
std_median = pd.Series(medians).std()

# 偏差値計算
for r in results:
    median = r["中央値"]
    if isinstance(median, int) and std_median != 0:
        r["中央値偏差値"] = round((median - mean_median) / std_median * 10 + 50, 2)
    else:
        r["中央値偏差値"] = None

# 列の順序
columns_order = [
    "取得日時", "タイトル", "ブランド", "中央値", "平均値", "最高点", "最低点", "中央値偏差値", "URL"
]

# Excelファイル削除（存在すれば）
if os.path.exists(output_file):
    try:
        os.remove(output_file)
    except Exception as e:
        print("⚠️ Excelファイルを閉じてから再実行してください。")
        raise e

# DataFrame出力
df = pd.DataFrame(results)[columns_order]
df.to_excel(output_file, index=False)

# 列幅調整
wb = load_workbook(output_file)
ws = wb.active

column_widths = {
    "A": 19, "B": 45, "C": 20, "D": 8, "E": 8,
    "F": 8, "G": 8, "H": 12, "I": 70
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 中央値の標準偏差を下部に表示
last_row = ws.max_row + 2
ws[f"A{last_row}"] = "中央値の標準偏差"
ws[f"B{last_row}"] = round(std_median, 2)

# 保存
wb.save(output_file)

print(f"\n📁 データ保存完了 → {output_file}")
